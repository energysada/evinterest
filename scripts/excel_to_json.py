#!/usr/bin/env python3
"""Convert evinterest.xlsx into JSON files for the static site."""

import json
import openpyxl
import sys
import os

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))
XLSX = os.path.join(PROJECT_ROOT, "fuel-prices-ev-interest", "ev interest tracker.xlsx")
OUT = os.path.join(PROJECT_ROOT, "data")

# Display order for indicators table (slide uses the same)
COUNTRY_ORDER = [
    "Australia", "New Zealand", "UK", "Germany", "Sweden", "France",
    "Italy", "Spain", "Denmark", "Norway",
    "US", "Vietnam", "Nepal", "Pakistan",
]


def parse_draft_table(wb):
    ws = wb["Draft Table"]

    # Row 2 has the Brent/edition info
    meta_text = ws.cell(row=2, column=1).value or ""
    title = ws.cell(row=1, column=1).value or ""

    # Find country headers (row 4, cols 3+)
    countries = []
    for col in range(3, 30):
        v = ws.cell(row=4, column=col).value
        if v:
            countries.append({"col": col, "name": v})
        else:
            break

    # Known section names
    section_names = {"WEEKLY VARIABLES", "MARKET CHARACTERISTICS"}

    # Parse metric rows
    sections = []
    current_section = None
    notes = []

    # Find all merged row ranges
    merged_rows = set()
    for m in ws.merged_cells.ranges:
        if m.min_col == 1 and m.max_col > 5:
            merged_rows.add(m.min_row)

    for row in range(5, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        label_str = str(label).strip() if label else ""

        # Helper: read country values for this row
        def read_values():
            values = []
            for c in countries:
                cell = ws.cell(row=row, column=c["col"])
                val = cell.value
                link = cell.hyperlink.target if cell.hyperlink else None
                values.append({
                    "country": c["name"],
                    "value": str(val) if val else "",
                    "link": link,
                })
            return values

        # Empty-label row = continuation of previous metric (if it has any country data)
        if not label_str:
            values = read_values()
            has_data = any(v["value"] for v in values)
            if has_data and current_section and current_section["metrics"]:
                # Append as a continuation row to the last metric
                current_section["metrics"].append({
                    "label": None,  # null label = continuation
                    "source": "",
                    "values": values,
                })
            continue

        # Section header — check by name (merge may have been lost)
        if label_str in section_names:
            current_section = {"name": label_str, "metrics": []}
            sections.append(current_section)
            continue

        # Merged row that's not a section = note
        if row in merged_rows:
            notes.append(label_str)
            continue

        # Skip notes that leaked in (long text, no country data)
        if len(label_str) > 80:
            notes.append(label_str)
            continue

        # Skip if no section started yet
        if current_section is None:
            continue

        # It's a metric row
        source = ws.cell(row=row, column=2).value or ""
        values = read_values()

        metric_label = label_str.replace("\n", " ")
        current_section["metrics"].append({
            "label": metric_label,
            "source": str(source).replace("\n", " ") if source else "",
            "values": values,
        })

    # Apply display order: pinned countries first, then remaining in Excel order
    excel_order = [c["name"] for c in countries]
    pinned = [c for c in COUNTRY_ORDER if c in excel_order]
    remaining = [c for c in excel_order if c not in pinned]
    display_order = pinned + remaining
    idx_by_name = {name: i for i, name in enumerate(display_order)}

    for section in sections:
        for metric in section["metrics"]:
            metric["values"].sort(key=lambda v: idx_by_name.get(v["country"], 999))

    return {
        "title": str(title),
        "subtitle": str(meta_text),
        "countries": display_order,
        "sections": sections,
        "notes": notes,
    }


def parse_news_feed(wb):
    ws = wb["Country News Feed"]

    headers = []
    for col in range(1, 10):
        v = ws.cell(row=1, column=col).value
        if v:
            headers.append(v)

    items = []
    for row in range(2, ws.max_row + 1):
        date_raw = ws.cell(row=row, column=1).value
        if not date_raw:
            continue
        # Normalize: datetime → date string YYYY-MM-DD
        if hasattr(date_raw, 'strftime'):
            date = date_raw.strftime('%Y-%m-%d')
        else:
            date = str(date_raw).split(' ')[0]

        # Get hyperlink from URL column (8)
        url_cell = ws.cell(row=row, column=8)
        url = url_cell.value or ""
        if url_cell.hyperlink:
            url = url_cell.hyperlink.target

        is_key = ws.cell(row=row, column=9).value == "Y"

        items.append({
            "date": str(date),
            "region": str(ws.cell(row=row, column=2).value or ""),
            "country": str(ws.cell(row=row, column=3).value or ""),
            "category": str(ws.cell(row=row, column=4).value or ""),
            "headline": str(ws.cell(row=row, column=5).value or ""),
            "data_point": str(ws.cell(row=row, column=6).value or ""),
            "source_name": str(ws.cell(row=row, column=7).value or ""),
            "source_url": str(url),
            "key_country": is_key,
        })

    return items


def main():
    if not os.path.exists(XLSX):
        print(f"Excel file not found: {XLSX}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX)

    # 1. Draft table
    tracker = parse_draft_table(wb)
    with open(os.path.join(OUT, "tracker.json"), "w") as f:
        json.dump(tracker, f, indent=2, ensure_ascii=False)
    print(f"tracker.json: {len(tracker['sections'])} sections, {len(tracker['countries'])} countries")

    # 2. News feed
    news = parse_news_feed(wb)
    with open(os.path.join(OUT, "news.json"), "w") as f:
        json.dump(news, f, indent=2, ensure_ascii=False)
    print(f"news.json: {len(news)} articles")

    # 3. Meta — Brent fetched live from Yahoo Finance (best-effort)
    from datetime import date
    today = date.today().isoformat()
    brent_str = "$94/bbl"
    brent_change = "+31% since Feb 28"
    try:
        import requests
        resp = requests.get(
            'https://query1.finance.yahoo.com/v8/finance/chart/BZ=F?interval=1d&range=2d',
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=10
        )
        data = resp.json()
        price = data['chart']['result'][0]['meta']['regularMarketPrice']
        baseline = 72
        pct = ((price - baseline) / baseline) * 100
        brent_str = f"${price:.0f}/bbl"
        brent_change = f"+{pct:.0f}% since Feb 28" if pct >= 0 else f"{pct:.0f}% since Feb 28"
        print(f"  Brent live: {brent_str} ({brent_change})")
    except Exception as e:
        print(f"  [WARN] Brent fetch failed, using fallback: {e}")
    meta = {
        "edition": 1,
        "date": today,
        "brent": brent_str,
        "brent_change": brent_change,
        "last_updated": today,
    }
    with open(os.path.join(OUT, "meta.json"), "w") as f:
        json.dump(meta, f, indent=2)
    print("meta.json written")

    # 4. Commentary (placeholder — Sada will edit)
    commentary = {
        "bullets": [
            "EV search interest has surged across all 9 tracked markets since the Iran conflict began on Feb 28, with Australia (+150%) and China (+128%) showing the strongest Google Trends response.",
            "Platform-level data confirms the trend: mobile.de (Germany) saw EV search share triple to 36%, AutoTrader UK reported EV leads up 28%, and Aramis Auto (France) saw EV sales share double to 12.7%.",
            "Brent crude hit $116/bbl after Houthi attacks on Israel on Mar 28, adding Red Sea disruption on top of the Hormuz closure. Oil executives warn Hormuz must reopen by mid-April.",
            "The interest signal is strongest in markets with high oil import dependence (>90%) and where governments have not cushioned pump prices — India, where excise cuts held prices flat, shows muted EV search response despite strong registration numbers.",
            "Used EVs are the fastest-moving segment: in the US, used EV sales rose 12% in Q1 while new EV sales fell 28%, and used EV average price ($34,821) is now within $1,300 of used gas cars.",
        ]
    }
    with open(os.path.join(OUT, "commentary.json"), "w") as f:
        json.dump(commentary, f, indent=2, ensure_ascii=False)
    print("commentary.json written")

    print("\nDone. All JSON files in data/")


if __name__ == "__main__":
    main()
