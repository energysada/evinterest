#!/usr/bin/env python3
"""Auto-update Google Trends and Brent indicators in the Draft Table.

Run daily as part of the ev-interest-country-scan pipeline.
Best-effort: logs warnings on failure but never crashes the pipeline.

Google Trends rows (13, 14, 15) get the full weekly progression format:
  W1% W2% W3% W4% W5% --> Current%
where each W is a 7-day average % change vs Feb 25-28 baseline.
"""

import os
import openpyxl
import requests
from datetime import date, datetime, timedelta

XLSX = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "fuel-prices-ev-interest", "ev interest tracker.xlsx",
)

# Pre-war baseline (matches gas price baseline)
BASELINE_START = "2026-02-25"
BASELINE_END = "2026-02-28"
BRENT_BASELINE = 72  # $/bbl pre-war
CRISIS_START = date(2026, 2, 28)

# Country geo codes
GEOS = {
    "US": "US", "UK": "GB", "Germany": "DE", "France": "FR",
    "Australia": "AU", "South Korea": "KR", "Vietnam": "VN", "India": "IN",
    "New Zealand": "NZ", "Sweden": "SE", "Italy": "IT", "Spain": "ES",
    "Denmark": "DK", "Norway": "NO",
    "Canada": "CA", "Singapore": "SG", "Thailand": "TH", "Philippines": "PH",
    "Japan": "JP", "Malaysia": "MY",
}

# Local-language equivalents
LOCAL_TERMS = {
    "electric car": {
        "DE": "Elektroauto", "FR": "voiture électrique",
        "KR": "전기차", "VN": "xe điện", "SE": "elbil",
        "IT": "auto elettrica", "ES": "coche eléctrico",
        "DK": "elbil", "NO": "elbil",
        "JP": "電気自動車", "TH": "รถยนต์ไฟฟ้า", "MY": "kereta elektrik",
    },
    "EV": {
        "DE": "E-Auto", "FR": "véhicule électrique",
        "KR": "EV", "VN": "ô tô điện", "SE": "EV",
        "IT": "EV", "ES": "EV", "DK": "EV", "NO": "EV",
        "JP": "EV", "TH": "EV", "MY": "EV",
    },
}

# Row mapping in Draft Table (after splitting Browsing → 3 rows, Used EVs → 3 rows, Showroom → 2 rows)
ROW_MAP = {
    "electric car": 12,
    "EV": 13,
    "used EV": 14,
}

USED_EV_COUNTRIES = {"US", "UK"}


def get_country_col(ws):
    countries = {}
    for c in range(3, 30):
        v = ws.cell(4, c).value
        if v:
            countries[v] = c
    return countries


def get_weekly_windows(today):
    """Build the weekly window dictionary based on today's date.

    Each weekly window is 7 days. Anchored to Feb 28 crisis start.
    Returns dict: baseline, w1, w2, ..., wN, current.
    """
    windows = {"baseline": (BASELINE_START, BASELINE_END)}
    week_n = 1
    week_start = CRISIS_START + timedelta(days=1)  # Mar 1
    while True:
        week_end = week_start + timedelta(days=6)
        if week_end >= today:
            windows["current"] = (week_start.isoformat(), today.isoformat())
            break
        windows[f"w{week_n}"] = (week_start.isoformat(), week_end.isoformat())
        week_start = week_end + timedelta(days=1)
        week_n += 1
        if week_n > 20:  # safety
            break
    return windows


def fetch_trend_progression(pytrends, term, geo, windows):
    """Returns dict of weekly avg values for the term."""
    try:
        end = date.today().isoformat()
        pytrends.build_payload(
            [term], cat=0,
            timeframe=f"2026-02-15 {end}",
            geo=geo,
        )
        data = pytrends.interest_over_time()
        if data.empty:
            return None
        data = data.reset_index()

        result = {}
        for k, (start, end_d) in windows.items():
            mask = (data['date'].astype(str) >= start) & (data['date'].astype(str) <= end_d)
            vals = data[mask][term].tolist()
            result[k] = sum(vals) / len(vals) if vals else 0
        return result
    except Exception as e:
        print(f"  [WARN] {term} {geo}: {e}")
        return None


def format_progression(weekly_avgs):
    """Format weekly_avgs into 'W1% W2% ... --> Current%'."""
    baseline = weekly_avgs.get("baseline", 0)
    if baseline == 0:
        return None

    week_keys = sorted(
        [k for k in weekly_avgs if k.startswith("w")],
        key=lambda x: int(x[1:])
    )

    parts = []
    for k in week_keys:
        v = weekly_avgs[k]
        p = ((v - baseline) / baseline) * 100
        parts.append(f"+{p:.0f}%" if p > 0 else f"{p:.0f}%" if p < 0 else "0%")

    cur = weekly_avgs.get("current", 0)
    if cur > 0:
        cur_p = ((cur - baseline) / baseline) * 100
        cur_str = f"+{cur_p:.0f}%" if cur_p > 0 else f"{cur_p:.0f}%" if cur_p < 0 else "0%"
    else:
        cur_str = "n/a"

    return " ".join(parts) + " --> " + cur_str


def update_google_trends(ws, countries):
    try:
        from pytrends.request import TrendReq
        import pandas as pd
        pd.set_option('future.no_silent_downcasting', True)
    except ImportError:
        print("  [WARN] pytrends not installed, skipping Google Trends update")
        return

    pytrends = TrendReq(hl='en-US', tz=300, timeout=(10, 25))
    today = date.today()
    windows = get_weekly_windows(today)
    print(f"  Weekly windows: {list(windows.keys())}")

    for term, row_num in ROW_MAP.items():
        # Update row label
        ws.cell(row_num, 1).value = f'Google: "{term}"\n(weekly % chg since Feb 28)'

        # Determine which countries to fetch for this term
        target_countries = USED_EV_COUNTRIES if term == "used EV" else GEOS.keys()

        for country_name in target_countries:
            geo = GEOS.get(country_name)
            if not geo:
                continue
            col = countries.get(country_name)
            if not col:
                continue

            search_term = LOCAL_TERMS.get(term, {}).get(geo, term)
            result = fetch_trend_progression(pytrends, search_term, geo, windows)
            if not result:
                continue

            formatted = format_progression(result)
            if formatted:
                ws.cell(row_num, col).value = formatted
                print(f"  GT [{term}] {country_name}: {formatted}")

    print("  Google Trends update complete")


def update_brent(ws):
    try:
        r = requests.get(
            'https://query1.finance.yahoo.com/v8/finance/chart/BZ=F?interval=1d&range=5d',
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=10
        )
        data = r.json()
        price = data['chart']['result'][0]['meta']['regularMarketPrice']
        pct = ((price - BRENT_BASELINE) / BRENT_BASELINE) * 100

        ws.cell(2, 1).value = f'Brent: ${price:.0f}/bbl (+{pct:.0f}% since Feb 28)  |  Edition #2'
        ws.cell(1, 1).value = f'Fuel Prices & EV Interest Tracker — Week of {date.today().strftime("%B %d, %Y")}'
        print(f"  Brent: ${price:.0f}/bbl (+{pct:.0f}%)")
        return price
    except Exception as e:
        print(f"  [WARN] Brent update failed: {e}")
        return None


def main():
    print("Updating indicators...")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['Draft Table']
    countries = get_country_col(ws)

    update_brent(ws)
    update_google_trends(ws, countries)

    wb.save(XLSX)
    print("Indicators saved.")


if __name__ == "__main__":
    main()
